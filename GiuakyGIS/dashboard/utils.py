from functools import wraps
from django.shortcuts import redirect
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth.models import User
from datetime import datetime
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from django.http import HttpResponse
import math


def admin_required(view_func):
    """Kiểm tra quyền admin"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.is_staff or request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        return redirect('/')
    return wrapper


def search_items(queryset, q, search_fields):
    if not q:
        return queryset
    query = Q()
    for field in search_fields:
        query |= Q(**{f"{field}__icontains": q})
    return queryset.filter(query)


def paginate_queryset(queryset, page, per_page=15):
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page)
    return page_obj, paginator


def get_stats():
    from dashboard.models import Store, Product, Order, Category

    stats = {
        'total_stores': Store.objects.count(),
        'total_products': Product.objects.count(),
        'total_orders': Order.objects.count(),
        'total_users': User.objects.count(),
        'total_categories': Category.objects.count(),
        'pending_orders': Order.objects.filter(status='Pending').count(),
        'processing_orders': Order.objects.filter(status='Processing').count(),
        'delivered_orders': Order.objects.filter(status='Delivered').count(),
    }
    return stats


# ==================== WAREHOUSE UTILITIES ====================

def generate_batch_number(batch_type):
    """Tạo số phiếu duy nhất cho phiếu nhập/xuất"""
    from dashboard.models import WarehouseBatch
    
    prefix = 'NK' if batch_type == 'import' else 'XK'
    today = datetime.now().strftime('%Y%m%d')
    
    # Lấy số lượng phiếu hôm nay
    count = WarehouseBatch.objects.filter(
        created_at__date=datetime.now().date(),
        batch_type=batch_type
    ).count() + 1
    
    return f"{prefix}{today}{count:04d}"


def search_stores_by_location(latitude, longitude, radius_km=5):
    """Tìm cửa hàng trong bán kính nhất định"""
    from dashboard.models import Store
    from django.db.models import FloatField
    from django.db.models.functions import ACos, Cos, Radians, Sin
    
    # Tính khoảng cách sử dụng công thức Haversine
    stores = Store.objects.all()
    nearby_stores = []
    
    for store in stores:
        distance = calculate_distance(latitude, longitude, store.latitude, store.longitude)
        if distance <= radius_km:
            nearby_stores.append({
                'store': store,
                'distance': round(distance, 2)
            })
    
    return sorted(nearby_stores, key=lambda x: x['distance'])


def calculate_distance(lat1, lon1, lat2, lon2):
    """Tính khoảng cách giữa hai điểm (Haversine formula)"""
    R = 6371  # Bán kính Trái Đất (km)
    
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    
    return R * c


def export_batch_to_excel(batch):
    """Xuất phiếu nhập/xuất kho ra file Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phiếu"
    
    # Tiêu đề
    ws['A1'] = f"Phiếu {batch.get_batch_type_display()}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    # Thông tin phiếu
    ws['A2'] = f"Số phiếu: {batch.batch_number}"
    ws['A3'] = f"Ngày: {batch.created_at.strftime('%d/%m/%Y')}"
    ws['A4'] = f"Kho: {batch.warehouse.store.name}"
    ws['A5'] = f"Nhà cung cấp: {batch.supplier}"
    
    # Tiêu đề bảng
    headers = ['STT', 'Sản phẩm', 'Số lượng', 'Đơn giá', 'Thành tiền']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Dữ liệu
    total = 0
    for idx, item in enumerate(batch.items.all(), 1):
        ws.cell(row=7+idx, column=1).value = idx
        ws.cell(row=7+idx, column=2).value = item.warehouse_item.product.name
        ws.cell(row=7+idx, column=3).value = item.quantity
        ws.cell(row=7+idx, column=4).value = item.unit_price
        ws.cell(row=7+idx, column=5).value = item.quantity * item.unit_price
        total += item.quantity * item.unit_price
    
    # Tổng tiền
    total_row = 8 + batch.items.count()
    ws.cell(row=total_row, column=4).value = "Tổng:"
    ws.cell(row=total_row, column=5).value = total
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    return wb


def export_batch_to_pdf(batch):
    """Xuất phiếu nhập/xuất kho ra file PDF (invoice)"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="phieu_{batch.batch_number}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Tiêu đề
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=20,
        alignment=1  # Center
    )
    story.append(Paragraph(f"Phiếu {batch.get_batch_type_display()}", title_style))
    story.append(Spacer(1, 12))
    
    # Thông tin phiếu
    info_data = [
        ['Số phiếu:', batch.batch_number],
        ['Ngày:', batch.created_at.strftime('%d/%m/%Y %H:%M')],
        ['Kho hàng:', batch.warehouse.store.name],
        ['Nhà cung cấp:', batch.supplier or 'N/A'],
        ['Người tạo:', batch.created_by.get_full_name() if batch.created_by else 'N/A'],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Bảng chi tiết
    data = [['STT', 'Sản phẩm', 'Số lượng', 'Đơn giá', 'Thành tiền']]
    total = 0
    
    for idx, item in enumerate(batch.items.all(), 1):
        subtotal = item.quantity * item.unit_price
        total += subtotal
        data.append([
            str(idx),
            item.warehouse_item.product.name,
            str(item.quantity),
            f"{item.unit_price:,.0f}",
            f"{subtotal:,.0f}"
        ])
    
    # Hàng tổng
    data.append(['', '', '', 'Tổng cộng:', f"{total:,.0f}"])
    
    table = Table(data, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')]),
    ]))
    story.append(table)
    
    doc.build(story)
    return response


def import_warehouse_from_excel(file_obj, warehouse, supplier=''):
    """Nhập dữ liệu kho từ file Excel"""
    try:
        import openpyxl
    except ImportError:
        return {'success': False, 'error': 'openpyxl chưa được cài đặt'}
    
    try:
        from dashboard.models import WarehouseBatch, WarehouseBatchItem, WarehouseItem, Product
        
        wb = openpyxl.load_workbook(file_obj)
        ws = wb.active
        
        # Tạo phiếu nhập kho mới
        batch = WarehouseBatch.objects.create(
            warehouse=warehouse,
            batch_type='import',
            batch_number=generate_batch_number('import'),
            supplier=supplier,
        )
        
        items_added = 0
        errors = []
        
        # Bỏ qua các hàng tiêu đề (giả sử từ hàng 2 trở đi)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row[0]:  # Bỏ qua hàng trống
                    continue
                
                product_name = row[0]
                quantity = int(row[1]) if row[1] else 0
                unit_price = float(row[2]) if row[2] else 0
                
                # Tìm sản phẩm
                try:
                    product = Product.objects.get(name=product_name)
                    warehouse_item, _ = WarehouseItem.objects.get_or_create(
                        warehouse=warehouse,
                        product=product
                    )
                    
                    # Thêm chi tiết phiếu
                    WarehouseBatchItem.objects.create(
                        batch=batch,
                        warehouse_item=warehouse_item,
                        quantity=quantity,
                        unit_price=unit_price
                    )
                    items_added += 1
                except Product.DoesNotExist:
                    errors.append(f"Hàng {row_idx}: Không tìm thấy sản phẩm '{product_name}'")
            except Exception as e:
                errors.append(f"Hàng {row_idx}: {str(e)}")
        
        batch.calculate_total()
        batch.save()
        
        return {
            'success': True,
            'batch_id': batch.id,
            'batch_number': batch.batch_number,
            'items_added': items_added,
            'errors': errors
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}
